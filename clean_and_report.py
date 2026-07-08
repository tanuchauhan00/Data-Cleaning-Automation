"""
Data Cleaning & Reporting Automation
-------------------------------------
Reads a raw (messy) sales CSV, cleans it, and produces:
  1. A cleaned CSV file
  2. A multi-sheet Excel report (raw data, cleaned data, summary, charts)
  3. PNG chart images

Usage:
    python clean_and_report.py <input_csv> <output_folder>

If no arguments are given, defaults to the sample dataset included in data/.
"""

import sys
import os
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.chart import BarChart, PieChart, Reference


# ---------------------------------------------------------------------------
# 1. LOAD
# ---------------------------------------------------------------------------
def load_data(path):
    df = pd.read_csv(path)
    return df


# ---------------------------------------------------------------------------
# 2. CLEAN
# ---------------------------------------------------------------------------
def clean_data(df):
    log = []
    original_rows = len(df)

    # Drop fully empty rows
    df = df.dropna(how="all")
    log.append(f"Removed {original_rows - len(df)} completely blank rows")

    # Remove exact duplicate rows
    before = len(df)
    df = df.drop_duplicates()
    log.append(f"Removed {before - len(df)} duplicate rows")

    # Standardize text columns: strip whitespace, fix casing
    for col in ["Region", "Product", "SalesRep"]:
        if col in df.columns:
            df[col] = df[col].astype(str).str.strip()
            df[col] = df[col].replace("nan", np.nan)

    if "Region" in df.columns:
        df["Region"] = df["Region"].str.title()
    if "Product" in df.columns:
        df["Product"] = df["Product"].str.title()

    # Standardize dates (multiple formats -> ISO)
    if "Date" in df.columns:
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce", format="mixed")
        n_bad_dates = df["Date"].isna().sum()
        log.append(f"Flagged {n_bad_dates} unparseable dates as missing")

    # Fix negative / invalid Units
    if "Units" in df.columns:
        bad_units = (df["Units"] < 0).sum()
        df.loc[df["Units"] < 0, "Units"] = np.nan
        log.append(f"Corrected {bad_units} negative Units values to missing")

    # Fill missing numeric values with column median (documented assumption)
    for col in ["Units", "UnitPrice"]:
        if col in df.columns:
            missing = df[col].isna().sum()
            median_val = df[col].median()
            df[col] = df[col].fillna(median_val)
            log.append(f"Filled {missing} missing '{col}' values with median ({median_val:.2f})")

    # Drop rows still missing a required key field (OrderID, Region, Product)
    before = len(df)
    df = df.dropna(subset=["OrderID", "Region", "Product"])
    log.append(f"Dropped {before - len(df)} rows missing key identifiers (OrderID/Region/Product)")

    # Recompute Revenue = Units * UnitPrice (was inconsistent/missing in raw data)
    df["Revenue"] = (df["Units"] * df["UnitPrice"]).round(2)

    df["OrderID"] = df["OrderID"].astype(int)
    df = df.sort_values("Date").reset_index(drop=True)

    return df, log


# ---------------------------------------------------------------------------
# 3. ANALYZE / SUMMARIZE
# ---------------------------------------------------------------------------
def summarize(df):
    summary = {
        "total_orders": len(df),
        "total_revenue": round(df["Revenue"].sum(), 2),
        "avg_order_value": round(df["Revenue"].mean(), 2),
        "date_range": f"{df['Date'].min().date()} to {df['Date'].max().date()}",
    }
    by_region = df.groupby("Region")["Revenue"].sum().sort_values(ascending=False).round(2)
    by_product = df.groupby("Product")["Revenue"].sum().sort_values(ascending=False).round(2)
    by_rep = df.groupby("SalesRep")["Revenue"].sum().sort_values(ascending=False).round(2)
    monthly = df.groupby(df["Date"].dt.to_period("M"))["Revenue"].sum().round(2)
    monthly.index = monthly.index.astype(str)
    return summary, by_region, by_product, by_rep, monthly


# ---------------------------------------------------------------------------
# 4. CHARTS (PNG, for quick viewing)
# ---------------------------------------------------------------------------
def make_charts(by_region, by_product, monthly, out_dir):
    plt.style.use("seaborn-v0_8-whitegrid")
    paths = {}

    fig, ax = plt.subplots(figsize=(6, 4))
    by_region.plot(kind="bar", ax=ax, color="#2E86AB")
    ax.set_title("Revenue by Region")
    ax.set_ylabel("Revenue ($)")
    plt.tight_layout()
    p = os.path.join(out_dir, "revenue_by_region.png")
    fig.savefig(p, dpi=150)
    plt.close(fig)
    paths["region"] = p

    fig, ax = plt.subplots(figsize=(6, 4))
    by_product.plot(kind="bar", ax=ax, color="#A23B72")
    ax.set_title("Revenue by Product")
    ax.set_ylabel("Revenue ($)")
    plt.tight_layout()
    p = os.path.join(out_dir, "revenue_by_product.png")
    fig.savefig(p, dpi=150)
    plt.close(fig)
    paths["product"] = p

    fig, ax = plt.subplots(figsize=(7, 4))
    monthly.plot(kind="line", marker="o", ax=ax, color="#F18F01")
    ax.set_title("Monthly Revenue Trend")
    ax.set_ylabel("Revenue ($)")
    plt.tight_layout()
    p = os.path.join(out_dir, "monthly_trend.png")
    fig.savefig(p, dpi=150)
    plt.close(fig)
    paths["monthly"] = p

    return paths


# ---------------------------------------------------------------------------
# 5. EXCEL REPORT
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial")
BODY_FONT = Font(name="Arial")


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def write_df_sheet(wb, name, df):
    ws = wb.create_sheet(name)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    style_header(ws, len(df.columns))
    for i, col in enumerate(df.columns, start=1):
        max_len = max(df[col].astype(str).map(lambda x: len(str(x))).max() if len(df) else 0, len(col)) + 2
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(max_len, 30)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
    return ws


def build_excel_report(raw_df, clean_df, log, summary, by_region, by_product, by_rep, monthly, chart_paths, out_path):
    wb = Workbook()
    wb.remove(wb.active)

    # --- Summary sheet ---
    ws = wb.create_sheet("Summary")
    ws["A1"] = "Data Cleaning & Reporting Automation - Summary"
    ws["A1"].font = Font(bold=True, size=14, name="Arial")
    ws.merge_cells("A1:D1")

    ws["A3"] = "Metric"
    ws["B3"] = "Value"
    style_header(ws, 2)
    rows = [
        ("Total Orders (cleaned)", summary["total_orders"]),
        ("Total Revenue", summary["total_revenue"]),
        ("Average Order Value", summary["avg_order_value"]),
        ("Date Range", summary["date_range"]),
        ("Raw Rows (before cleaning)", len(raw_df)),
        ("Rows Removed / Corrected", len(raw_df) - summary["total_orders"]),
    ]
    for i, (k, v) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=k).font = BODY_FONT
        ws.cell(row=i, column=2, value=v).font = BODY_FONT
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    log_row = i + 3
    ws.cell(row=log_row, column=1, value="Cleaning Steps Applied:").font = Font(bold=True, name="Arial")
    for j, line in enumerate(log, start=log_row + 1):
        ws.cell(row=j, column=1, value=f"- {line}").font = BODY_FONT

    img_row = j + 3
    for k, key in enumerate(["region", "product", "monthly"]):
        img = XLImage(chart_paths[key])
        img.width, img.height = 380, 260
        ws.add_image(img, f"A{img_row + k * 15}")

    # --- Breakdown sheets ---
    write_df_sheet(wb, "By Region", by_region.reset_index().rename(columns={"Revenue": "Total Revenue"}))
    write_df_sheet(wb, "By Product", by_product.reset_index().rename(columns={"Revenue": "Total Revenue"}))
    write_df_sheet(wb, "By Sales Rep", by_rep.reset_index().rename(columns={"Revenue": "Total Revenue"}))
    write_df_sheet(wb, "Monthly Trend", monthly.reset_index().rename(columns={"Date": "Month", "Revenue": "Total Revenue"}))

    # --- Native Excel chart (region) built from data on the sheet, so it's dynamic ---
    region_ws = wb["By Region"]
    chart = BarChart()
    chart.title = "Revenue by Region"
    chart.y_axis.title = "Revenue ($)"
    data = Reference(region_ws, min_col=2, min_row=1, max_row=region_ws.max_row)
    cats = Reference(region_ws, min_col=1, min_row=2, max_row=region_ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    region_ws.add_chart(chart, "D2")

    # --- Cleaned & Raw data sheets ---
    write_df_sheet(wb, "Cleaned Data", clean_df)
    write_df_sheet(wb, "Raw Data (original)", raw_df)

    wb.save(out_path)


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main():
    base = os.path.dirname(os.path.abspath(__file__))
    input_csv = sys.argv[1] if len(sys.argv) > 1 else os.path.join(base, "data", "raw_sales_data.csv")
    out_dir = sys.argv[2] if len(sys.argv) > 2 else os.path.join(base, "output")
    os.makedirs(out_dir, exist_ok=True)

    raw_df = load_data(input_csv)
    clean_df, log = clean_data(raw_df.copy())
    summary, by_region, by_product, by_rep, monthly = summarize(clean_df)
    chart_paths = make_charts(by_region, by_product, monthly, out_dir)

    clean_csv_path = os.path.join(out_dir, "cleaned_sales_data.csv")
    clean_df.to_csv(clean_csv_path, index=False)

    excel_path = os.path.join(out_dir, "Sales_Data_Report.xlsx")
    build_excel_report(raw_df, clean_df, log, summary, by_region, by_product, by_rep, monthly, chart_paths, excel_path)

    print("Cleaning log:")
    for line in log:
        print(" -", line)
    print(f"\nCleaned CSV saved to: {clean_csv_path}")
    print(f"Excel report saved to: {excel_path}")


if __name__ == "__main__":
    main()
